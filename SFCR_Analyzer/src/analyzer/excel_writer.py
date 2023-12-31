import os
import re
import pdfplumber
import pandas as pd
import numpy as np
from openpyxl import load_workbook

def write_to_excel(filename, merged_tables):
    excel_default = "DEF_BILANS-global.xlsx"
    excel_path = filename + '-global.xlsx'
    pattern = r'[A-Z][0-9]{4}'
    workbook = load_workbook(excel_default)
    specific_key = '02.01.02'

    print(f'Excel path: {excel_path}')

    # Iterate through the rows and cells
    if specific_key in merged_tables and f'S.{specific_key}' in workbook.sheetnames:
        merged_table = merged_tables[specific_key]
        if not merged_table.empty:
            sheet = workbook[f'S.{specific_key}']
            print(f'Processing sheet: S.{sheet}')

            # Iterate through each row in the DataFrame
            for _, row in merged_table.iterrows():
                for i, cell in enumerate(row):
                    if pd.notna(cell) and re.match(pattern, str(cell)):
                        # The next value in the row is the value to write to Excel
                        value_to_write = row[i + 1] if i + 1 < len(row) else None

                        # Find the cell in the Excel sheet that matches the pattern
                        for excel_row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, min_col=1,
                                                         max_col=sheet.max_column):
                            for excel_cell in excel_row:
                                if excel_cell.value == cell:
                                    # Write the value to the next cell (column + 1)
                                    next_cell = sheet.cell(row=excel_cell.row, column=excel_cell.column + 1)
                                    next_cell.value = value_to_write
                                    break

    # Save the workbook
    workbook.save(excel_path)
    workbook.close()
